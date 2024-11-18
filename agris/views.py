from django.shortcuts import render, redirect
from django.contrib.auth.forms import UserCreationForm
from .forms import RegistrationForm
from django.http import JsonResponse
from .models import NPKSensor, AmbientSensor, SoilSensor, Settings
from django.utils import timezone
from django.core.serializers import serialize
from django.utils.timezone import localtime
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy


def admin_dashboard(request):
    # Fetch the current status of the fetching setting
    setting, created = Settings.objects.get_or_create(name='fetching_enabled', defaults={'value': False})
    context = {'fetching_enabled': setting.value}
    return render(request, "admin/adminDashboard.html", context)

def toggle_fetching(request):
    if request.method == 'POST':
        setting, created = Settings.objects.get_or_create(name='fetching_enabled', defaults={'value': True})
        setting.value = not setting.value  # Toggle the value
        setting.save()
        return JsonResponse({'success': True, 'fetching_enabled': setting.value})
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')  # Redirect to login page after successful registration
    else:
        form = RegistrationForm()
    return render(request, 'register.html', {'form': form})


class CustomLoginView(LoginView):
    def get_success_url(self):
        # Check user type or profile data to decide redirection
        if self.request.user.is_staff:
            return reverse_lazy('dashboard')  # Redirect staff to admin dashboard
        return reverse_lazy('dashboard')  # Redirect regular users to user dashboard
    
def dashboard(request):
    return render(request, 'main/dashboard.html')

def graph1(request):
    return render(request, 'main/frame/graph1.html')

def graph2(request):
    return render(request, 'main/frame/graph2.html')

def graph3(request):
    return render(request, 'main/frame/graph3.html')

def map(request):
    return render(request, 'main/frame/map.html')

import requests
from django.http import JsonResponse
from .models import NPKSensor


def is_fetching_enabled():
    try:
        setting = Settings.objects.get(name='fetching_enabled')
        return setting.value
    except Settings.DoesNotExist:
        return True  # Default to enabled if not set
def save_sensor_data(request):
    if not is_fetching_enabled():
        return JsonResponse({"success": False, "error": "Fetching is disabled."})
    # API endpoints for sensor data
    api_urls = {
        "NITROGEN": "https://sgp1.blynk.cloud/external/api/get?token=KAg_k8iVJioh6oIh16TUvoZsVjcyVpRG&v6",
        "PHOSPHORUS": "https://sgp1.blynk.cloud/external/api/get?token=KAg_k8iVJioh6oIh16TUvoZsVjcyVpRG&v7",
        "POTASSIUM": "https://sgp1.blynk.cloud/external/api/get?token=KAg_k8iVJioh6oIh16TUvoZsVjcyVpRG&v8",
    }

    try:
        # Fetch data from APIs
        nitrogen = float(requests.get(api_urls["NITROGEN"]).json())
        phosphorus = float(requests.get(api_urls["PHOSPHORUS"]).json())
        potassium = float(requests.get(api_urls["POTASSIUM"]).json())

        # Save to database
        NPKSensor.objects.create(
            nitrogen=nitrogen,
            phosphorus=phosphorus,
            potassium=potassium,
        )

        return JsonResponse({"success": True, "message": "Data saved successfully."})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

def fetch_sensor_data(request):
    if not is_fetching_enabled():
        return JsonResponse({"success": False, "error": "Fetching is disabled."})
    # Fetch the latest sensor data
    sensor_data = NPKSensor.objects.all().order_by('-timestamp')  # Get the last 10 records
    data = [
        {
            'timestamp': localtime(sd.timestamp).strftime('%H:%M:%S'),  # Convert to local time and format timestamp
            'nitrogen': sd.nitrogen,
            'phosphorus': sd.phosphorus,
            'potassium': sd.potassium,
        }
        for sd in sensor_data
    ]
    return JsonResponse({'success': True, 'data': data})

def save_ambient_sensor_data(request):
    if not is_fetching_enabled():
        return JsonResponse({"success": False, "error": "Fetching is disabled."})
    # API endpoints for ambient sensor data
    api_urls = {
        "A_HUMIDITY": "https://sgp1.blynk.cloud/external/api/get?token=KAg_k8iVJioh6oIh16TUvoZsVjcyVpRG&v0",
        "A_TEMPERATURE": "https://sgp1.blynk.cloud/external/api/get?token=KAg_k8iVJioh6oIh16TUvoZsVjcyVpRG&v1",
    }

    try:
        # Fetch data from APIs
        A_Humid = float(requests.get(api_urls["A_HUMIDITY"]).json())
        A_Temp = float(requests.get(api_urls["A_TEMPERATURE"]).json())

        # Save to database
        AmbientSensor.objects.create(
            A_Humid=A_Humid,
            A_Temp=A_Temp,
        )

        return JsonResponse({"success": True, "message": "Ambient data saved successfully."})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

def fetch_ambient_sensor_data(request):
    if not is_fetching_enabled():
        return JsonResponse({"success": False, "error": "Fetching is disabled."})
    # Fetch the latest ambient sensor data
    sensor_data = AmbientSensor.objects.all().order_by('-timestamp')  # Get the last 10 records
    data = [
        {
            'timestamp': localtime(sd.timestamp).strftime('%H:%M:%S'),  # Convert to local time and format timestamp
            'A_Humid': sd.A_Humid,
            'A_Temp': sd.A_Temp,
        }
        for sd in sensor_data
    ]
    return JsonResponse({'success': True, 'data': data})

def save_soil_sensor_data(request):
    if not is_fetching_enabled():
        return JsonResponse({"success": False, "error": "Fetching is disabled."})
    # API endpoints for soil sensor data
    api_urls = {
        "SOIL_HUMIDITY": "https://sgp1.blynk.cloud/external/api/get?token=KAg_k8iVJioh6oIh16TUvoZsVjcyVpRG&v2",
        "SOIL_TEMPERATURE": "https://sgp1.blynk.cloud/external/api/get?token=KAg_k8iVJioh6oIh16TUvoZsVjcyVpRG&v3",
    }

    try:
        # Fetch data from APIs
        S_Humid = float(requests.get(api_urls["SOIL_HUMIDITY"]).json())
        S_Temp = float(requests.get(api_urls["SOIL_TEMPERATURE"]).json())

        # Save to database
        SoilSensor.objects.create(
            S_Humid=S_Humid,
            S_Temp=S_Temp,
        )

        return JsonResponse({"success": True, "message": "Soil data saved successfully."})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

def fetch_soil_sensor_data(request):
    if not is_fetching_enabled():
        return JsonResponse({"success": False, "error": "Fetching is disabled."})
    # Fetch the latest soil sensor data
    sensor_data = SoilSensor.objects.all().order_by('-timestamp')
    data = [
        {
            'timestamp': localtime(sd.timestamp).strftime('%H:%M:%S'),  # Convert to local time and format timestamp
            'S_Humid': sd.S_Humid,
            'S_Temp': sd.S_Temp,
        }
        for sd in sensor_data
    ]
    return JsonResponse({'success': True, 'data': data})

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils.timezone import localtime
from .models import NPKSensor, AmbientSensor, SoilSensor

def generate_report(request):
    # Create an in-memory output file for the new workbook
    output = BytesIO()

    # Create a workbook and add a worksheet for each sensor
    wb = Workbook()

    # NPK Sensor Data
    ws1 = wb.active
    ws1.title = "NPK Sensor Data"
    ws1.append(["Timestamp", "Nitrogen", "Phosphorus", "Potassium"])
    for data in NPKSensor.objects.all():
        ws1.append([
            localtime(data.timestamp).replace(tzinfo=None),  # Remove timezone
            data.nitrogen,
            data.phosphorus,
            data.potassium,
        ])
    
    # Ambient Sensor Data
    ws2 = wb.create_sheet(title="Ambient Sensor Data")
    ws2.append(["Timestamp", "Temperature", "Humidity"])
    for data in AmbientSensor.objects.all():
        ws2.append([
            localtime(data.timestamp).replace(tzinfo=None),  # Remove timezone
            data.A_Temp,
            data.A_Humid,
        ])
    
    # Soil Sensor Data
    ws3 = wb.create_sheet(title="Soil Sensor Data")
    ws3.append(["Timestamp", "Temperature", "Humidity"])
    for data in SoilSensor.objects.all():
        ws3.append([
            localtime(data.timestamp).replace(tzinfo=None),  # Remove timezone
            data.S_Temp,
            data.S_Humid,
        ])
    
    # Save the workbook to the output stream
    wb.save(output)
    output.seek(0)

    # Send the file as an HTTP response
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="SensorDataReport.xlsx"'
    
    return response

from django.shortcuts import render
from django.core.paginator import Paginator
from .models import NPKSensor, AmbientSensor, SoilSensor, Settings
import requests

def report(request):
    # Get data from models
    npk_data = NPKSensor.objects.all().order_by('-timestamp')
    ambient_data = AmbientSensor.objects.all().order_by('-timestamp')
    soil_data = SoilSensor.objects.all().order_by('-timestamp')

    # Pagination setup
    npk_paginator = Paginator(npk_data, 10)  # 10 items per page
    ambient_paginator = Paginator(ambient_data, 10)
    soil_paginator = Paginator(soil_data, 10)

    npk_page = request.GET.get('npk_page')
    ambient_page = request.GET.get('ambient_page')
    soil_page = request.GET.get('soil_page')

    npk_data = npk_paginator.get_page(npk_page)
    ambient_data = ambient_paginator.get_page(ambient_page)
    soil_data = soil_paginator.get_page(soil_page)

    # Settings
    settings_data = Settings.objects.all()

    # Check API connection status
    api_url = "https://sgp1.blynk.cloud/external/api/get?token=KAg_k8iVJioh6oIh16TUvoZsVjcyVpRG&v10"
    try:
        response = requests.get(api_url, timeout=5)
        api_status = response.status_code == 200
    except requests.RequestException:
        api_status = False

    context = {
        'npk_data': npk_data,
        'ambient_data': ambient_data,
        'soil_data': soil_data,
        'settings_data': settings_data,
        'api_status': api_status,
    }

    return render(request, 'main/frame/report.html', context)
